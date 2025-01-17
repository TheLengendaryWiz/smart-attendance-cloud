import json
import openpyxl
import random


def name(name):
    with open("data.json", "r") as f:
        cont = f.read()
        cont = json.loads(cont)
        final = []
        for i in cont.keys():
            if name.lower() in cont[i]["name"].lower():
                new = cont[i]
                new["usn"] = i
                final.append(new)

    return json.dumps(final)


def usn(usn):
    with open("data.json", "r") as f:
        cont = f.read()
        print(cont)
        cont = json.loads(cont)
    try:
        print("from drv" + json.dumps(cont[usn.upper()]))
        cont[usn.upper()]["usn"] = list(cont.keys())[list(cont.values()).index(cont[usn.upper()])]
        return json.dumps(cont[usn.upper()])
    except KeyError:
        return ""


def clas(classec):
    clas = classec.split(" ")[0]
    sec = classec.split(" ")[1]

    with open("data.json", "r") as f:
        cont = f.read()

        cont = json.loads(cont)
        final = []
        print("from class" + str(final))
        for i in cont.keys():

            if str(cont[i]["clas"]) == str(clas) and str(cont[i]["section"]).lower() == sec.lower():
                new = cont[i]
                new["usn"] = i
                final.append(new)
        print("from class" + str(final))
    return json.dumps(final)


def gExClass(classec, date):
    students = json.loads(clas(classec))
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = classec
    ws["B1"] = "Attendance"
    ws["C1"] = "Entry Time (yr:m:d:hr:min:sec)"
    ws["D1"] = "Exit Time (yr:m:d:hr:min:sec)"
    for i in students:
        print("hello")
        entry = []
        for j in i["entries"]:
            if date in j["start"]:
                try:
                    entry = [j["start"], j["end"]]
                except KeyError:
                    entry = [j["start"], ""]
        if len(entry) > 0:
            attendance = "present"
            ws.append([f"Name: {i['name']} Usn: {i['usn']}", attendance, entry[0], entry[1]])
        else:
            attendance = "absent"
            ws.append([f"Name: {i['name']} Usn: {i['usn']}", attendance, "", ""])

        #print([f"Name: {i['name']} Usn: {i['usn']}", attendance, entry[0], entry[1]])

        

    return wb
